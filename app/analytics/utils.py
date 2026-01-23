from io import BytesIO

from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from sqlalchemy import func

from app import db
from app.models import Order, OrderItem, SMMStats
from app.utils.db import get_usd_uah_rate


def build_monthly_stats_days(start_date_str: str, end_date_str: str):
    if not start_date_str or not end_date_str:
        raise ValueError("startDate/endDate required")

    start_dt = datetime.strptime(start_date_str, "%Y-%m-%d").replace(hour=0, minute=0, second=0)
    end_dt = datetime.strptime(end_date_str, "%Y-%m-%d").replace(hour=23, minute=59, second=59)

    order_data = (
        db.session.query(
            func.date(Order.created_at).label("day"),
            func.count(func.distinct(Order.id)).label("order_count"),
            func.sum(OrderItem.quantity * OrderItem.unit_price).label("total_sales"),
            func.sum(OrderItem.quantity * OrderItem.unit_margin).label("total_margin"),
        )
        .join(OrderItem, Order.id == OrderItem.order_id)
        .filter(Order.created_at >= start_dt, Order.created_at <= end_dt)
        .group_by(func.date(Order.created_at))
        .order_by(func.date(Order.created_at))
        .all()
    )

    orders_by_day = {
        datetime.strptime(record.day, "%Y-%m-%d").date(): {
            "order_count": record.order_count,
            "total_sales": float(record.total_sales or 0),
            "total_margin": float(record.total_margin or 0),
        }
        for record in order_data
    }

    smm_data = (
        db.session.query(SMMStats)
        .filter(SMMStats.date >= start_dt.date(), SMMStats.date <= end_dt.date())
        .all()
    )
    smm_by_day = {sd.date: sd for sd in smm_data}

    days = []
    day = start_dt

    while day <= end_dt:
        order_info = orders_by_day.get(
            day.date(), {"order_count": 0, "total_sales": 0.0, "total_margin": 0.0}
        )

        smm_info = smm_by_day.get(day.date())

        if smm_info and not smm_info.usd_rate:
            smm_info.usd_rate = get_usd_uah_rate(day)
            db.session.commit()

        spends_usd = float(smm_info.spends) if smm_info else 0.0
        rate = float(smm_info.usd_rate) if (smm_info and smm_info.usd_rate) else 0.0
        spends_uah = spends_usd * rate

        days.append(
            {
                "date": day.date().isoformat(),
                "order_count": order_info["order_count"],
                "total_sales": order_info["total_sales"],
                "total_margin": order_info["total_margin"],
                "smm_spends_usd": spends_usd,
                "smm_spends_uah": spends_uah,
                "smm_coverage": smm_info.coverage if smm_info else 0,
                "smm_clicks": smm_info.clicks if smm_info else 0,
                "smm_direct_messages": smm_info.direct_messages if smm_info else 0,
                "revenue": order_info["total_margin"] - spends_uah,
            }
        )

        day += timedelta(days=1)

    return days


def build_excel_data(days):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Stats"

    headers = [
        ("Date", "date"),
        ("Orders", "order_count"),
        ("Total Sales", "total_sales"),
        ("Total Margin", "total_margin"),
        ("SMM Spends (USD)", "smm_spends_usd"),
        ("SMM Spends (UAH)", "smm_spends_uah"),
        ("Coverage", "smm_coverage"),
        ("Clicks", "smm_clicks"),
        ("Direct Messages", "smm_direct_messages"),
        ("Revenue", "revenue"),
    ]

    # Write header row
    ws.append([h[0] for h in headers])
    header_font = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Excel number formats (space thousands separator)
    money_fmt = "# ##0.00"
    int_fmt = "0"
    date_fmt = "yyyy-mm-dd"

    # Write data rows
    for d in days:
        # parse ISO date string to actual date object for Excel date type
        dt = datetime.strptime(d["date"], "%Y-%m-%d").date()

        ws.append([
            dt,
            int(d.get("order_count", 0) or 0),
            float(d.get("total_sales", 0.0) or 0.0),
            float(d.get("total_margin", 0.0) or 0.0),
            float(d.get("smm_spends_usd", 0.0) or 0.0),
            float(d.get("smm_spends_uah", 0.0) or 0.0),
            int(d.get("smm_coverage", 0) or 0),
            int(d.get("smm_clicks", 0) or 0),
            int(d.get("smm_direct_messages", 0) or 0),
            float(d.get("revenue", 0.0) or 0.0),
        ])

        r = ws.max_row
        ws.cell(row=r, column=1).number_format = date_fmt

        # Integers
        for c in (2, 7, 8, 9):
            ws.cell(row=r, column=c).number_format = int_fmt

        # Money/decimal columns
        for c in (3, 4, 5, 6, 10):
            ws.cell(row=r, column=c).number_format = money_fmt

    # Reasonable column widths
    widths = [12, 8, 14, 14, 16, 16, 10, 8, 16, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(output)
    output.seek(0)
    return output